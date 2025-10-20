import random
from fpdf import FPDF
from datetime import datetime
import os
import platform
import subprocess
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


def open_file(filename):
    system = platform.system()
    try:
        if system == "Darwin":  # macOS
            subprocess.run(["open", filename])
        elif system == "Windows":
            os.startfile(filename)
        elif system == "Linux":
            subprocess.run(["xdg-open", filename])
        else:
            print(f"Cannot auto-open file on {system}.")
    except Exception as e:
        print(f"Failed to open file: {e}")

def generate_realistic_mark():
    mark = int(random.gauss(mu=65, sigma=15))  # Mean 65, Std Dev 15
    return max(0, min(mark, 100))  # Clamp between 0 and 100

class Student:
    def __init__(self,name,reg_no,roll_no,school,stream):
        self.name = name
        self.reg_no = reg_no
        self.roll_no = roll_no
        self.school = school
        self.stream = stream

class Marksheet :
    def __init__(self,student, subjects ):
        self.student = student
        self.subjects = subjects
        self.marks = {sub: generate_realistic_mark() for sub in subjects}
        self.lowest_mark = None
        self.best5_total = self.calculate_best5_total()
        self.percent = self.calculate_best5_total () / 5
        self.grade = self.get_overall_grade()

    def get_grade (self, mark):
        if mark >= 90 :
            return 'A+'
        elif mark >= 80 :
            return 'A'
        elif mark >= 70 :
            return 'B+'
        elif mark >= 60 :
            return 'B'
        elif mark >= 50 :
            return 'C'
        elif mark >= 40 :
            return 'D'
        else :
            return 'F'
    
    def get_overall_grade(self):
    # Count how many subjects have grade 'F'
        failed_subjects = sum(1 for mark in self.marks.values() if self.get_grade(mark) == 'F')

        if failed_subjects > 1:
            return 'F'  #  Failed due to multiple subject failures

    # Otherwise, grade based on percentage
        if self.percent >= 90:
            return 'A+'
        elif self.percent >= 80:
            return 'A'
        elif self.percent >= 70:
            return 'B+'
        elif self.percent >= 60:
            return 'B'
        elif self.percent >= 50:
            return 'C'
        elif self.percent >= 40:
            return 'D'
        else:
            return 'F'

        
    def calculate_best5_total(self):
        subject_count = len(self.subjects)

        if subject_count > 6:
            raise ValueError("Number of subjects cannot be more than 6")

        if subject_count == 6:
            # Drop the subject with the lowest mark
            lowest = min(self.marks, key=self.marks.get)
            self.lowest_mark = lowest
            return sum(self.marks.values()) - self.marks[lowest]

        # For 5 or fewer subjects, count all
        return sum(self.marks.values())

        
    def generate_pdf(self, filename = None):
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", size = 12)

        pdf.cell(200, 10, txt="West Bengal Council of Higher Secondary Education", ln=True, align='C')
        pdf.cell(200, 10, txt="MARKSHEET", ln=True, align='C')
        pdf.ln(10)

        pdf.cell(200, 10, txt=f"Date: {datetime.now().strftime('%d-%m-%Y')}", ln=True)
        pdf.cell(200, 10, txt=f"Name: {self.student.name.upper()}", ln=True)
        pdf.cell(200, 10, txt=f"Registration No: {self.student.reg_no}", ln=True)
        pdf.cell(200, 10, txt=f"Roll No: {self.student.roll_no}", ln=True)
        pdf.cell(200, 10, txt=f"School: {self.student.school.upper()}", ln=True)
        pdf.cell(200, 10, txt=f"Stream: {self.student.stream}", ln=True)
        pdf.ln(10)

        pdf.set_font("Arial", 'B', 12)
        pdf.cell(65, 10, "Subject", border=1, align='C')
        pdf.cell(65, 10, "Marks", border=1, align='C')
        pdf.cell(65, 10, "Grade", border=1, ln=True, align='C')
        pdf.set_font("Arial", size=12)

        for sub, mark in self.marks.items():
            grade = self.get_grade(mark)
            if grade == "F":
                pdf.set_text_color(255, 0, 0)  #  Red for failed
            else:
                pdf.set_text_color(0, 0, 0)    #  Black for others
            pdf.cell(65, 10, sub, border=1, align='C')
            pdf.cell(65, 10, str(mark), border=1, align='C')
            pdf.cell(65, 10, grade, border=1, ln=True, align='C')
        pdf.set_text_color(0, 0, 0)  # Reset to black
        pdf.ln(10)
        pdf.cell(200, 10, txt=f"Best 5 Total: {self.best5_total} / 500", ln=True)
        pdf.cell(200, 10, txt=f"Percentage: {self.percent:.2f}%", ln=True)
        pdf.cell(200, 10, txt=f"Overall Grade: {self.grade}", ln=True)

        if not filename :
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"{self.student.name.replace(' ', '_')}_{timestamp}.pdf"
        pdf.output(filename)
        open_file(filename)

def export_all_students_to_single_sheet(students_data, filename="All_Results.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "All Students"

    # Header row
    headers = [
        "Name", "Reg No", "Roll No", "School", "Stream",
        "Subject", "Marks", "Grade", "Best 5 Total", "Percentage", "Overall Grade","Status"
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    pass_count = 0
    fail_count = 0
    total_students = len(students_data)

    # Add each student's data
    for marksheet in students_data:
        student = marksheet.student
        failed_subjects = sum(1 for mark in marksheet.marks.values() if marksheet.get_grade(mark) == 'F')
        status = "Fail" if failed_subjects > 1 else "Pass"
        if status == "Fail":
            fail_count += 1
        else:
            pass_count += 1

        for sub, mark in marksheet.marks.items():
            grade = marksheet.get_grade(mark)
            row = [
                student.name,
                student.reg_no,
                student.roll_no,
                student.school,
                student.stream,
                sub,
                mark,
                grade,
                marksheet.best5_total,
                f"{marksheet.percent:.2f}%",
                marksheet.grade,
                status
            ]
            ws.append(row)

            if status == "Fail":
                for cell in ws[ws.max_row]:
                    cell.fill = red_fill
            else:
                for cell in ws[ws.max_row]:
                    cell.fill = green_fill


    wb.save(filename)
    ws.append([])  # Blank row
    ws.append(["Summary", "", "", "", "", "", "", "", "", "", "", ""])
    ws.append(["Total Students", total_students])
    ws.append(["Total Passed", pass_count])
    ws.append(["Total Failed", fail_count])

    # Calculate percentages
    pass_percent = (pass_count / total_students) * 100 if total_students else 0
    fail_percent = (fail_count / total_students) * 100 if total_students else 0

    ws.append(["Pass Percentage", f"{pass_percent:.2f}%"])
    ws.append(["Fail Percentage", f"{fail_percent:.2f}%"])

    # Save again to include summary
    wb.save(filename)
    print(f"Excel file saved as '{filename}'")


def main():
    students_data = []
    num_students = int(input("How many students do you want to enter? "))

    for i in range(num_students):
        print(f"\n--- Enter details for Student {i+1} ---")
        name = input("Enter Student Name: ")
        reg_no = input("Enter Registration Number: ")
        roll_no = input("Enter Roll Number: ")
        school = input("Enter School Name: ")

        print("\nSelect Stream:\n1. Science\n2. Arts\n3. Commerce")
        stream_choice = int(input("Enter choice: "))
        stream = ["Science", "Arts", "Commerce"][stream_choice - 1]

        num_subjects = int(input("How many subjects? (5 or 6): "))
        subjects = [input(f"Enter name of subject {j+1}: ").upper() for j in range(num_subjects)]

        student = Student(name, reg_no, roll_no, school, stream)
        marksheet = Marksheet(student, subjects)
        marksheet.generate_pdf()
        students_data.append(marksheet)

    export_all_students_to_single_sheet(students_data)
    print("\nAll marksheets generated and Excel file saved successfully!")

    
if __name__ == "__main__" :
    main() 